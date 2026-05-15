#!/usr/bin/env python3
"""
review_matcher.py v2.0 — Enhanced Student Review Intelligence Module

Matches a scholar against a structured student-review database and produces a JSON
summary with:
  - Rating statistics and 14-dimension sentiment breakdowns
  - Credibility scoring per review
  - Structured investigation leads with:
      • Severity computed dynamically (mentions × sentiment × credibility)
      • Concrete evidence quotes extracted from original reviews
      • Cross-dimensional consistency analysis
      • Verification action items with specific databases and queries

Usage:
    python review_matcher.py --db "../reviews.xlsx" --name "CASE_014" --school "重庆邮电大学" --output ./data/reviews.json
"""

import json
import re
import sys
import argparse
from pathlib import Path
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from typing import Optional


# ---------------------------------------------------------------------------
# 1. Loader
# ---------------------------------------------------------------------------

def load_xlsx(db_path: str) -> list[dict]:
    """Load rows from xlsx. Pandas is preferred because it handles merged cells."""
    try:
        import pandas as pd
        df = pd.read_excel(db_path, sheet_name=0)
        df = df.ffill()
        return df.to_dict(orient="records")
    except ImportError:
        pass
    try:
        import openpyxl
        wb = openpyxl.load_workbook(db_path, data_only=False)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        last_values = {}
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            row = {}
            for h, v in zip(headers, raw_row):
                if h is None:
                    continue
                if v is not None:
                    last_values[h] = v
                    row[h] = v
                else:
                    row[h] = last_values.get(h)
            rows.append(row)
        return rows
    except ImportError:
        raise RuntimeError("Cannot read xlsx. Please install pandas or openpyxl.")


def extract_reviews(row: dict) -> list[tuple[str, float | None]]:
    """Return list of (review_text, rating) for a row."""
    texts = []
    rating = row.get("评分")
    try:
        rating = float(rating)
    except (ValueError, TypeError):
        rating = None

    primary_keys = [k for k in row.keys() if "评价" in k or "评论" in k or "内容" in k]
    for k in primary_keys:
        val = row.get(k)
        if val not in (None, "", "-", " "):
            texts.append((str(val).strip(), rating))

    for k, v in row.items():
        if v not in (None, "", "-", " ") and isinstance(v, str) and len(v) > 20:
            if k not in primary_keys:
                texts.append((v.strip(), rating))
    return texts


# ---------------------------------------------------------------------------
# 2. Dimension parser (14 dimensions)
# ---------------------------------------------------------------------------

def _dim_pat(name: str, aliases: str, others: str) -> tuple:
    pat = rf"(?:^|\n)\s*(?:[一二三四五六七八九十\d]+[、.．\s]+)?(?:{aliases})[：:]\s*(.*?)(?=\n(?:{others})[：:]|$)"
    return (name, pat)

_OTHER_DIMS = (
    "导师辨识特征|学术水平|科研经费|学生补助|师生关系|工作时间|学生前途|"
    "自证认识导师|自我介绍|自证|研究方向|毕业要求|毕业条件|发文要求|论文排名|论文署名|"
    "组会|周会|例会|指导方式|人品|性格|为人|行为及人品|"
    "实习|就业|读博|出国|是否推荐|推荐|不建议|后悔|庆幸|避雷|"
    "实验室氛围|团队氛围|课题组氛围"
)

DIMENSION_PATTERNS = [
    _dim_pat("导师辨识特征", "导师辨识特征", _OTHER_DIMS),
    _dim_pat("学术水平", "学术水平", _OTHER_DIMS),
    _dim_pat("科研经费", "科研经费", _OTHER_DIMS),
    _dim_pat("学生补助", "学生补助", _OTHER_DIMS),
    _dim_pat("师生关系", "师生关系", _OTHER_DIMS),
    _dim_pat("工作时间", "工作时间", _OTHER_DIMS),
    _dim_pat("学生前途", "学生前途", _OTHER_DIMS),
    _dim_pat("自证认识导师", "自证认识导师|自我介绍|自证", _OTHER_DIMS),
    _dim_pat("毕业要求与论文署名", "毕业要求|毕业条件|发文要求|论文排名|论文署名", _OTHER_DIMS),
    _dim_pat("组会与指导方式", "组会|周会|例会|指导方式", _OTHER_DIMS),
    _dim_pat("人品与性格", "人品|性格|为人|行为及人品", _OTHER_DIMS),
    _dim_pat("实习与就业支持", "实习|就业|读博|出国", _OTHER_DIMS),
    _dim_pat("推荐意愿", "是否推荐|推荐|不建议|后悔|庆幸|避雷", _OTHER_DIMS),
    _dim_pat("实验室氛围", "实验室氛围|团队氛围|课题组氛围", _OTHER_DIMS),
]


def parse_dimensions(text: str) -> dict:
    dimensions = {}
    for dim_name, pattern in DIMENSION_PATTERNS:
        m = re.search(pattern, text, re.DOTALL)
        if m:
            dimensions[dim_name] = m.group(1).strip()
    return dimensions


# ---------------------------------------------------------------------------
# 3. Sentiment analysis (enhanced with intensity)
# ---------------------------------------------------------------------------

POSITIVE_WORDS = [
    "好", "不错", "优秀", "棒", "强", "高", "推荐", "值得", "满意", "愉快",
    "尊重", "支持", "关心", "负责", "认真", "充实", "受益", "顺利", "轻松",
    "自由", "大方", "厚道", "公正", "公平", "靠谱", "nice", "good", "great",
    "佩服", "欣赏", "认可", "赞同", "欣慰", "开心", "幸福", "温暖", "感动",
    "牛", "厉害", "大牛", "牛导",
]

NEGATIVE_WORDS = [
    "差", "烂", "坑", "垃圾", "恶心", "压榨", "剥削", "无耻", "卑鄙", "阴险",
    "虚伪", "自私", "抠门", "苛刻", "暴躁", "骂人", "威胁", "恐吓", "PUA",
    "pua", "后悔", "不建议", "避雷", "逃离", "快跑", "别来", "惨", "痛苦",
    "抑郁", "焦虑", "迷茫", "放养", "不管", "冷漠", "卡毕业", "延毕", "抢一作",
    "套钱", "撒谎", "欺骗", "刻薄", "无视", "辱骂", "侮辱", "贬低", "打压",
    "排挤", "孤立", "猜忌", "多疑", "控制欲", "偷窥", "窥探", "骚扰", "性骚扰",
    "猥亵", "暴怒", "易怒", "神经质", "变态", "人渣", "败类", "黑心", "恶毒",
    "吝啬", "独断", "霸道", "蛮横", "不讲理", "吹毛求疵", "斤斤计较", "睚眦必报",
    "过河拆桥", "表里不一", "两面三刀", "阳奉阴违", "欺上瞒下", "招摇撞骗",
    "刚愎自用", "目中无人", "飞扬跋扈", "自以为是", "固执己见", "冥顽不灵",
    "油盐不进", "独断专行", "一手遮天", "为所欲为", "肆无忌惮", "明目张胆",
    "恬不知耻", "厚颜无耻", "恩将仇报", "忘恩负义", "落井下石", "趁火打劫",
    "趋炎附势", "阿谀奉承", "溜须拍马", "奴颜婢膝", "卑躬屈膝", "低三下四",
    "忍气吞声", "委曲求全", "逆来顺受", "任人宰割", "出气筒", "替罪羊",
    "背锅侠", "冤大头", "和稀泥", "骑墙派", "墙头草", "两面派", "双面人",
    "伪君子", "真小人", "笑面虎", "铁公鸡", "周扒皮", "黄世仁", "严监生", "葛朗台",
    "奴役", "奴隶", "畜生", "禽兽", "狗", "猪", "废物", "蠢货", "脑残", "弱智",
    "奇葩", "神经病", "疯子", "傻子", "呆子", "白痴",
]

INTENSE_NEGATIVE_WORDS = [
    "性骚扰", "猥亵", "强奸", "性侵", "暴力", "打人", "殴打", "动手",
    "自杀", "跳楼", "割腕", "自残", "死亡", "逼死", "逼疯",
]


def analyze_sentiment(text: str) -> dict:
    """Return detailed sentiment analysis with intensity score."""
    pos = sum(text.count(w) for w in POSITIVE_WORDS)
    neg = sum(text.count(w) for w in NEGATIVE_WORDS)
    intense = sum(text.count(w) for w in INTENSE_NEGATIVE_WORDS)

    if intense > 0:
        label = "intense_negative"
    elif neg > pos * 1.5:
        label = "strong_negative"
    elif neg > pos:
        label = "negative"
    elif pos > neg:
        label = "positive"
    else:
        label = "neutral"

    # Intensity score: 0-100
    total_sentiment_words = pos + neg + intense * 3
    text_len = len(text)
    if text_len > 0:
        density = min(total_sentiment_words / (text_len / 50), 1.0)  # normalize
    else:
        density = 0

    intensity = int((neg * 10 + intense * 50 + density * 20) / (pos + 1))
    intensity = min(intensity, 100)

    return {
        "label": label,
        "positive_count": pos,
        "negative_count": neg,
        "intense_negative_count": intense,
        "intensity_score": intensity,
    }


# ---------------------------------------------------------------------------
# 4. Credibility scoring (enhanced)
# ---------------------------------------------------------------------------

def compute_credibility(text: str, dimensions: dict) -> float:
    score = 0.0
    if "自证认识导师" in dimensions or "自我介绍" in dimensions:
        score += 0.30
    if len(text) >= 300:
        score += 0.25
    elif len(text) >= 150:
        score += 0.15
    elif len(text) >= 50:
        score += 0.05

    dim_count = len(dimensions)
    if dim_count >= 10:
        score += 0.30
    elif dim_count >= 7:
        score += 0.20
    elif dim_count >= 4:
        score += 0.10
    elif dim_count >= 2:
        score += 0.05

    if re.search(r"\d{4}|\d+篇|\d+块|\d+元|\d+%|\d+年|\d+月|\d+点|\d+:\d+", text):
        score += 0.15
    if re.search(r"比如|例如|有一次|某次|具体|直接|亲眼|亲身经历|亲耳听到", text):
        score += 0.10

    # Bonus for self-contradiction disclosure (shows awareness)
    if any(w in text for w in ["保证", "实话", "真实", "亲身经历", "亲眼"]):
        score += 0.05

    return round(min(score, 1.0), 2)


# ---------------------------------------------------------------------------
# 5. Investigation Leads (v2.0 — enhanced)
# ---------------------------------------------------------------------------

@dataclass
class LeadDef:
    id: str
    label: str
    keywords: list[str]
    base_severity: str  # low / medium / high / critical
    verify_action: str
    verify_databases: list[str]
    related_dimensions: list[str]
    cross_threshold: int = 2  # min number of related dimensions with negative sentiment to upgrade severity


LEAD_DEFINITIONS = [
    LeadDef(
        id="delayed_graduation",
        label="延期毕业指控",
        keywords=["延毕", "卡毕业", "不让毕业", "拖毕业", "毕业难", "延期", "推迟毕业", "卡答辩", "卡论文"],
        base_severity="high",
        verify_action="核对该导师名下学生实际毕业年份与标准学制差异",
        verify_databases=["CNKI学位论文库: 检索导师=XXX 的博硕士论文，记录毕业年份", "学校官网: 查看历届毕业生名单"],
        related_dimensions=["毕业要求与论文署名", "师生关系", "学生前途"],
        cross_threshold=2,
    ),
    LeadDef(
        id="authorship_extraction",
        label="署名榨取指控",
        keywords=["抢一作", "导师一作", "抢论文", "署名", "文章被抢", "被抢一作", "霸占一作"],
        base_severity="high",
        verify_action="统计近三年论文，核对导师一作率与学生一作率",
        verify_databases=["CNKI/WoS: 检索导师近3年论文，统计一作/通讯分布", "学校学位论文库: 核对学生学位论文章节发表署名"],
        related_dimensions=["毕业要求与论文署名", "师生关系", "学术水平"],
        cross_threshold=2,
    ),
    LeadDef(
        id="funding_embezzlement",
        label="经费克扣/套现",
        keywords=["套钱", "套现", "克扣经费", "挪用", "经费私用", "报销", "虚假发票", "虚假报销"],
        base_severity="high",
        verify_action="查NSFC/省级基金在研项目经费使用情况，结合实验室规模判断",
        verify_databases=["NSFC大数据门户: 查询在研项目及经费", "学校财务公示: 查看大型设备采购记录", "全国哲学社会科学工作办公室: 查社科项目"],
        related_dimensions=["科研经费", "学生补助", "师生关系"],
        cross_threshold=2,
    ),
    LeadDef(
        id="stipend_theft",
        label="学生补助克扣",
        keywords=["补助少", "不发工资", "克扣补助", "200块", "低保", "侮辱", "这点钱", "返钱", "退还"],
        base_severity="medium",
        verify_action="核实学校规定的学生补助标准与实际发放记录",
        verify_databases=["学校研究生院官网: 查看补助标准文件", "学生银行流水(如有): 核对实际到账金额"],
        related_dimensions=["学生补助", "师生关系", "科研经费"],
        cross_threshold=2,
    ),
    LeadDef(
        id="internship_banned",
        label="限制实习",
        keywords=["不允许实习", "不让实习", "禁止实习", "实习不让", "不放实习", "卡实习"],
        base_severity="medium",
        verify_action="通过LinkedIn/ResearchGate追踪往届学生职业轨迹，核实实习经历",
        verify_databases=["LinkedIn: 搜索毕业生 profile，查看实习记录", "知乎/小红书: 搜索'XXX导师 实习'"],
        related_dimensions=["实习与就业支持", "学生前途", "工作时间"],
        cross_threshold=2,
    ),
    LeadDef(
        id="high_workload",
        label="高强度工作制度",
        keywords=["打卡", "996", "加班", "晚上", "周末", "早8", "晚10", "考勤", "门禁"],
        base_severity="medium",
        verify_action="核实实验室工作时长要求，与学校规定对比",
        verify_databases=["学校研究生院: 查看是否有超时长工作投诉记录", "实地走访(如有条件): 观察实验室作息"],
        related_dimensions=["工作时间", "学生补助", "师生关系"],
        cross_threshold=1,
    ),
    LeadDef(
        id="toxic_culture",
        label="负面实验室文化",
        keywords=["骂人", "PUA", "压榨", "恶心", "威胁", "龌龊", "恶心", "虚伪", "两面派", "帮派", "站队"],
        base_severity="high",
        verify_action="增加peer consultation权重，定向核实具体事件；关注是否有集体投诉",
        verify_databases=["学校纪委/研究生院: 查看是否有学生投诉记录(需申请)", "导师评价网/小红书/知乎: 搜索导师名字+负面关键词"],
        related_dimensions=["人品与性格", "实验室氛围", "师生关系"],
        cross_threshold=2,
    ),
    LeadDef(
        id="mental_abuse",
        label="精神虐待/PUA",
        keywords=["PUA", "精神控制", "洗脑", "打击", "否定", "摧毁自信", "人格侮辱", "羞辱", "贬低"],
        base_severity="high",
        verify_action="与多位在读/已毕业学生私下交流，核实是否存在系统性精神打压",
        verify_databases=["研学网/导师评价网: 查看是否有类似描述", "知乎/小红书: 搜索'XXX导师 PUA'"],
        related_dimensions=["人品与性格", "师生关系", "实验室氛围"],
        cross_threshold=2,
    ),
    LeadDef(
        id="verbal_abuse",
        label="言语辱骂",
        keywords=["骂人", "骂学生", "脏话", "辱骂", "人身攻击", "侮辱", "恶语", "讽刺", "挖苦"],
        base_severity="high",
        verify_action="核实辱骂频率和场景，区分偶尔情绪激动与习惯性辱骂",
        verify_databases=["同上(mental_abuse)", "微信聊天记录(如有): 收集证据"],
        related_dimensions=["人品与性格", "师生关系", "组会与指导方式"],
        cross_threshold=1,
    ),
    LeadDef(
        id="career_sabotage",
        label="职业破坏",
        keywords=["卡找工作", "不让找工作", "耽误找工作", "影响就业", "找不到工作", "故意"],
        base_severity="high",
        verify_action="追踪毕业生职业轨迹，核实是否存在系统性就业困难",
        verify_databases=["LinkedIn: 统计毕业生去向和行业分布", "脉脉/看准网: 搜索实验室毕业生评价"],
        related_dimensions=["学生前途", "实习与就业支持", "师生关系"],
        cross_threshold=2,
    ),
    LeadDef(
        id="absent_shepherd",
        label="放养型管理",
        keywords=["放养", "不管学生", "见不到人", "找不到人", "不指导", "无指导", "自生自灭"],
        base_severity="medium",
        verify_action="与近三年论文署名模式交叉验证（通讯作者占比、产出连续性）",
        verify_databases=["CNKI: 核对学生论文是否有导师署名", "学校学位论文库: 查看学生论文致谢部分"],
        related_dimensions=["组会与指导方式", "学术水平", "师生关系"],
        cross_threshold=1,
    ),
    LeadDef(
        id="graduation_barrier",
        label="毕业门槛异常",
        keywords=["毕业要求", "发文要求", "必须发", "卡论文", "SCI", "CSSCI", "核心期刊", "额外要求"],
        base_severity="medium",
        verify_action="对照学院官方培养方案，核实毕业要求是否超出常规标准",
        verify_databases=["学校研究生院官网: 下载培养方案PDF", "学院通知公告: 查看是否有额外毕业要求文件"],
        related_dimensions=["毕业要求与论文署名", "学术水平", "学生前途"],
        cross_threshold=1,
    ),
    LeadDef(
        id="academic_isolation",
        label="学术孤立",
        keywords=["孤立", "排挤", "不让学生参加", "不让开会", "不让交流", "封锁信息", "闭门造车"],
        base_severity="medium",
        verify_action="核对学生是否有参加学术会议、交流项目的记录",
        verify_databases=["学校国际合作处: 查询学生出国/参会记录", "会议官网: 搜索学生姓名+会议名称"],
        related_dimensions=["组会与指导方式", "学术水平", "实验室氛围"],
        cross_threshold=2,
    ),
    LeadDef(
        id="nepotism",
        label="任人唯亲/帮派文化",
        keywords=["任人唯亲", "偏袒", "关系户", "嫡系", "非嫡系", "区别对待", "不公平", "帮派", "小团体"],
        base_severity="medium",
        verify_action="观察资源分配（经费、论文署名、推荐机会）是否存在明显偏向",
        verify_databases=["论文数据库: 统计不同学生的署名机会和期刊级别", "访谈多位学生: 了解资源分配实际情况"],
        related_dimensions=["师生关系", "实验室氛围", "学生前途"],
        cross_threshold=2,
    ),
    LeadDef(
        id="research_obsolescence",
        label="研究方向过时",
        keywords=["过时", "啃老本", "炒冷饭", "没创新", "落后", "脱节", "老旧", "跟不上", "不前沿"],
        base_severity="medium",
        verify_action="核对其近5年论文的主题分布和引用前沿文献比例",
        verify_databases=["CNKI/WoS: 分析论文关键词演变", "Google Scholar: 查看近3年高引文献是否被引用"],
        related_dimensions=["学术水平", "研究方向", "学生前途"],
        cross_threshold=1,
    ),
    LeadDef(
        id="data_fabrication_pressure",
        label="逼迫数据造假",
        keywords=["造假", "改数据", "编数据", "凑数据", "要结果", "必须显著", "P值", "逼"],
        base_severity="critical",
        verify_action="核查论文中的统计数据一致性，必要时使用统计检验工具",
        verify_databases=["pdfplumber+Python: 提取论文表格进行统计一致性审查", "原始数据请求: 向作者或期刊索取原始数据"],
        related_dimensions=["学术水平", "科研经费", "人品与性格"],
        cross_threshold=1,
    ),
    LeadDef(
        id="sexual_harassment",
        label="性骚扰",
        keywords=["性骚扰", "猥亵", "动手动脚", "占便宜", "揩油", "骚扰", "不当接触", "暧昧"],
        base_severity="critical",
        verify_action="极度敏感，必须谨慎处理。如有多个独立来源提及，建议直接建议当事人向学校纪委或公安机关反映",
        verify_databases=["学校纪委: 查询投诉记录(需正规渠道)", "公安机关: 如有受害者愿意报案", "注意: 严禁散布未经证实的性侵指控"],
        related_dimensions=["人品与性格", "师生关系", "实验室氛围"],
        cross_threshold=1,
    ),
    LeadDef(
        id="physical_abuse",
        label="肢体暴力",
        keywords=["打人", "动手", "暴力", "推搡", "殴打", "体罚", "砸东西"],
        base_severity="critical",
        verify_action="同sexual_harassment。如有多个独立来源，建议立即向学校保卫处和公安机关报告",
        verify_databases=["学校保卫处: 查询报案记录", "公安机关: 如有受害者愿意报案"],
        related_dimensions=["人品与性格", "师生关系", "实验室氛围"],
        cross_threshold=1,
    ),
]

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}
SEVERITY_UPGRADE = {"low": "medium", "medium": "high", "high": "critical", "critical": "critical"}


def extract_evidence_quotes(text: str, keywords: list[str], context: int = 40) -> list[str]:
    """Extract sentences containing any of the keywords, with surrounding context."""
    quotes = []
    for kw in keywords:
        for m in re.finditer(re.escape(kw), text):
            start = max(m.start() - context, 0)
            end = min(m.end() + context, len(text))
            quote = text[start:end].strip()
            # Clean up: don't break in the middle of a word
            if start > 0 and text[start - 1] not in "\n。！？；":
                # Find previous sentence boundary
                prev_boundary = max(text.rfind("\n", 0, start), text.rfind("。", 0, start),
                                    text.rfind("！", 0, start), text.rfind("？", 0, start))
                if prev_boundary != -1:
                    start = prev_boundary + 1
            if end < len(text) and text[end] not in "\n。！？；":
                next_boundary = min(
                    (text.find("\n", end) if text.find("\n", end) != -1 else len(text)),
                    (text.find("。", end) if text.find("。", end) != -1 else len(text)),
                    (text.find("！", end) if text.find("！", end) != -1 else len(text)),
                    (text.find("？", end) if text.find("？", end) != -1 else len(text)),
                )
                end = next_boundary
            quote = text[start:end].strip()
            if quote and quote not in quotes:
                quotes.append(quote)
    return quotes[:5]  # max 5 quotes per lead


def compute_dynamic_severity(
    lead_def: LeadDef,
    mention_count: int,
    review_items: list[tuple[str, dict]],
    dim_summary: dict,
    scored_reviews: list[dict],
) -> tuple[str, list[str]]:
    """Compute dynamic severity with upgrade rules. Returns (severity, reasons)."""
    severity = lead_def.base_severity
    reasons = [f"基础级别: {lead_def.base_severity}"]

    # Upgrade 1: mention count
    if mention_count >= 5:
        severity = SEVERITY_UPGRADE[severity]
        reasons.append(f"提及次数多({mention_count}次): 级别上调")
    elif mention_count >= 3:
        if severity in ("low", "medium"):
            severity = SEVERITY_UPGRADE[severity]
            reasons.append(f"提及次数较多({mention_count}次): 级别上调")

    # Upgrade 2: cross-dimensional consistency
    negative_dim_count = 0
    for dim in lead_def.related_dimensions:
        if dim in dim_summary:
            dist = dim_summary[dim].get("sentiment_distribution", {})
            total = sum(dist.values())
            if total > 0:
                neg_ratio = (dist.get("negative", 0) + dist.get("strong_negative", 0) + dist.get("intense_negative", 0)) / total
                if neg_ratio >= 0.5:
                    negative_dim_count += 1
    if negative_dim_count >= lead_def.cross_threshold:
        severity = SEVERITY_UPGRADE[severity]
        reasons.append(f"交叉维度验证通过({negative_dim_count}/{len(lead_def.related_dimensions)}个相关维度呈负面): 级别上调")

    # Upgrade 3: credibility of reviews mentioning this lead
    avg_cred = sum(r["credibility_score"] for r in scored_reviews) / len(scored_reviews) if scored_reviews else 0
    if avg_cred >= 0.6:
        if severity in ("low", "medium"):
            severity = SEVERITY_UPGRADE[severity]
            reasons.append(f"高可信度评价支撑(平均可信度{avg_cred:.2f}): 级别上调")

    # Upgrade 4: intense negative words present
    combined_text = " ".join(text for text, _ in review_items)
    intense_count = sum(combined_text.count(w) for w in INTENSE_NEGATIVE_WORDS)
    if intense_count > 0 and severity != "critical":
        severity = SEVERITY_UPGRADE[severity]
        reasons.append(f"出现极端负面词汇({intense_count}次): 级别上调")

    return severity, reasons


def generate_leads(
    review_items: list[tuple[str, dict]],
    dim_summary: dict,
    scored_reviews: list[dict],
) -> list[dict]:
    """Generate structured investigation leads with evidence quotes and dynamic severity."""
    all_texts = [text for text, _ in review_items]
    combined = "\n".join(all_texts)
    leads = []

    for lead_def in LEAD_DEFINITIONS:
        # Count mentions across all reviews
        mention_count = 0
        affected_reviews = []
        all_quotes = []

        for i, (text, dims) in enumerate(review_items):
            text_mentions = sum(text.count(kw) for kw in lead_def.keywords)
            if text_mentions > 0:
                mention_count += text_mentions
                affected_reviews.append(i)
                quotes = extract_evidence_quotes(text, lead_def.keywords)
                for q in quotes:
                    if q not in all_quotes:
                        all_quotes.append(q)

        if mention_count == 0:
            continue

        # Compute dynamic severity
        severity, severity_reasons = compute_dynamic_severity(
            lead_def, mention_count, review_items, dim_summary, scored_reviews
        )

        # Affected review credibility
        affected_creds = [scored_reviews[i]["credibility_score"] for i in affected_reviews if i < len(scored_reviews)]
        avg_affected_cred = round(sum(affected_creds) / len(affected_creds), 2) if affected_creds else 0

        # Cross-dimensional negative ratio
        neg_dim_count = 0
        for dim in lead_def.related_dimensions:
            if dim in dim_summary:
                dist = dim_summary[dim].get("sentiment_distribution", {})
                total = sum(dist.values())
                if total > 0:
                    neg = dist.get("negative", 0) + dist.get("strong_negative", 0) + dist.get("intense_negative", 0)
                    if neg / total >= 0.5:
                        neg_dim_count += 1

        leads.append({
            "id": lead_def.id,
            "label": lead_def.label,
            "mention_count": mention_count,
            "affected_reviews": len(affected_reviews),
            "severity": severity,
            "severity_reasons": severity_reasons,
            "base_severity": lead_def.base_severity,
            "confidence_level": "L3" if severity in ("high", "critical") else ("L2" if severity == "medium" else "L2"),
            "verify_action": lead_def.verify_action,
            "verify_databases": lead_def.verify_databases,
            "related_dimensions": lead_def.related_dimensions,
            "cross_validation": {
                "negative_dimensions": neg_dim_count,
                "total_related": len(lead_def.related_dimensions),
                "ratio": round(neg_dim_count / len(lead_def.related_dimensions), 2) if lead_def.related_dimensions else 0,
            },
            "evidence_quotes": all_quotes[:3],  # top 3 unique quotes
            "avg_credibility_of_affected": avg_affected_cred,
        })

    # Sort by severity (critical first), then by mention count (desc)
    leads.sort(key=lambda x: (SEVERITY_ORDER.get(x["severity"], 99), -x["mention_count"]))
    return leads


# ---------------------------------------------------------------------------
# 6. Aggregators (enhanced with sentiment details)
# ---------------------------------------------------------------------------

def compute_dimension_summary(review_items: list[tuple[str, dict]]) -> dict:
    """Aggregate parsed dimensions with detailed sentiment across all review items."""
    all_dims = defaultdict(list)
    for text, dims in review_items:
        for dim_name, dim_text in dims.items():
            sentiment = analyze_sentiment(dim_text)
            all_dims[dim_name].append({
                "text": dim_text,
                "sentiment": sentiment,
            })

    summary = {}
    for dim_name, entries in all_dims.items():
        sentiments = Counter(e["sentiment"]["label"] for e in entries)
        avg_intensity = sum(e["sentiment"]["intensity_score"] for e in entries) / len(entries)
        summary[dim_name] = {
            "mention_count": len(entries),
            "sentiment_distribution": dict(sentiments),
            "dominant_sentiment": sentiments.most_common(1)[0][0],
            "average_intensity": round(avg_intensity, 1),
            "sample_quotes": [e["text"] for e in entries[:3]],
        }
    return summary


def build_radar_data(summary: dict) -> list[dict]:
    """Prepare radar-chart friendly data from dimension summary."""
    radar = []
    sentiment_score = {"positive": 5, "neutral": 3, "negative": 2, "strong_negative": 1, "intense_negative": 0}
    for dim_name, data in summary.items():
        dist = data.get("sentiment_distribution", {})
        total = sum(dist.values())
        if total == 0:
            continue
        weighted = sum(sentiment_score.get(s, 3) * c for s, c in dist.items()) / total
        radar.append({
            "dimension": dim_name,
            "score": round(weighted, 2),
            "max": 5,
            "mention_count": data["mention_count"],
            "dominant_sentiment": data["dominant_sentiment"],
        })
    order = {d[0]: i for i, d in enumerate(DIMENSION_PATTERNS)}
    radar.sort(key=lambda x: order.get(x["dimension"], 99))
    return radar


# ---------------------------------------------------------------------------
# 7. Cross-dimensional anomaly detection
# ---------------------------------------------------------------------------

def detect_anomaly_patterns(summary: dict) -> list[dict]:
    """Detect unusual patterns across dimensions that may indicate systematic issues."""
    anomalies = []

    # Pattern 1: All dimensions negative except "导师辨识特征" (fake positive front)
    non_id_dims = [d for d in summary.keys() if d != "导师辨识特征"]
    if non_id_dims:
        neg_count = sum(1 for d in non_id_dims if summary[d]["dominant_sentiment"] in ("negative", "strong_negative", "intense_negative"))
        if neg_count / len(non_id_dims) >= 0.7:
            id_sent = summary.get("导师辨识特征", {}).get("dominant_sentiment", "neutral")
            if id_sent in ("positive", "neutral"):
                anomalies.append({
                    "pattern": "表面包装/实质负面",
                    "description": "多数维度呈负面，但'导师辨识特征'维度偏正面，可能存在'表面包装'与'实质问题'的反差",
                    "severity": "medium",
                    "affected_dimensions": [d for d in non_id_dims if summary[d]["dominant_sentiment"].startswith("negative")],
                })

    # Pattern 2: Extreme polarization (some very positive, some very negative)
    pos_dims = [d for d in summary if summary[d]["dominant_sentiment"] == "positive"]
    neg_dims = [d for d in summary if summary[d]["dominant_sentiment"] in ("negative", "strong_negative", "intense_negative")]
    if len(pos_dims) >= 2 and len(neg_dims) >= 2:
        anomalies.append({
            "pattern": "极端两极分化",
            "description": f"{len(pos_dims)}个维度正面、{len(neg_dims)}个维度负面，评价存在严重分歧，可能反映'选择性优待'或'水军刷评'",
            "severity": "medium",
            "positive_dimensions": pos_dims,
            "negative_dimensions": neg_dims,
        })

    # Pattern 3: "学生前途" strongly negative while "学术水平" neutral/positive
    future = summary.get("学生前途", {})
    academic = summary.get("学术水平", {})
    if future.get("dominant_sentiment", "").startswith("negative") and academic.get("dominant_sentiment", "") in ("positive", "neutral"):
        anomalies.append({
            "pattern": "学术尚可但就业差",
            "description": "学术水平评价中性或正面，但学生前途评价负面，可能存在'重学术轻就业'或'故意耽误学生求职'的问题",
            "severity": "high",
            "affected_dimensions": ["学生前途", "实习与就业支持"],
        })

    # Pattern 4: All work-related dimensions negative (systemic exploitation)
    work_dims = ["工作时间", "学生补助", "师生关系", "实验室氛围"]
    work_neg = sum(1 for d in work_dims if d in summary and summary[d]["dominant_sentiment"].startswith("negative"))
    if work_neg >= 3:
        anomalies.append({
            "pattern": "系统性工作环境问题",
            "description": f"{work_neg}/4个工作环境相关维度呈负面，反映可能存在系统性的学生待遇问题",
            "severity": "high",
            "affected_dimensions": [d for d in work_dims if d in summary and summary[d]["dominant_sentiment"].startswith("negative")],
        })

    return anomalies


# ---------------------------------------------------------------------------
# 8. Matching
# ---------------------------------------------------------------------------

def match_scholar(rows: list[dict], name: str, school: str | None, college: str | None) -> list[dict]:
    matched = []
    for row in rows:
        row_name = str(row.get("姓名", "")).strip()
        row_school = str(row.get("学校", "")).strip()
        row_college = str(row.get("学院", "")).strip()
        if row_name != name:
            continue
        if school and row_school != school:
            continue
        if college and row_college != college:
            continue
        matched.append(row)
    return matched


# ---------------------------------------------------------------------------
# 9. Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Match scholar against student review database (v2.0)")
    parser.add_argument("--db", "-d", required=True, help="Path to xlsx review database")
    parser.add_argument("--name", "-n", required=True, help="Scholar name (exact match)")
    parser.add_argument("--school", "-s", help="School/University name (exact match)")
    parser.add_argument("--college", "-c", help="College/Department name (exact match)")
    parser.add_argument("--output", "-o", required=True, help="Path to output JSON")
    parser.add_argument("--top-reviews", "-t", type=int, default=5, help="Number of top-credibility reviews to include")
    args = parser.parse_args()

    print(f"[INFO] Loading database: {args.db}")
    rows = load_xlsx(args.db)
    print(f"[INFO] Total rows loaded: {len(rows)}")

    print(f"[INFO] Matching: name='{args.name}', school='{args.school or '*'}', college='{args.college or '*'}'")
    matched = match_scholar(rows, args.name, args.school, args.college)
    print(f"[INFO] Matched rows: {len(matched)}")

    if not matched:
        result = {
            "matched": False,
            "name": args.name,
            "school": args.school,
            "college": args.college,
            "review_count": 0,
            "message": "No matching records found in the review database.",
        }
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"[INFO] No matches. Empty result saved to: {args.output}")
        return

    ratings = []
    all_review_texts = []
    review_items = []

    for row in matched:
        rating = row.get("评分")
        try:
            ratings.append(float(rating))
        except (ValueError, TypeError):
            pass
        for text, _ in extract_reviews(row):
            all_review_texts.append(text)
            dims = parse_dimensions(text)
            review_items.append((text, dims))

    # Credibility scoring per review
    scored_reviews = []
    for text, dims in review_items:
        sentiment = analyze_sentiment(text)
        scored_reviews.append({
            "text_preview": text[:400],
            "credibility_score": compute_credibility(text, dims),
            "dimensions_parsed": list(dims.keys()),
            "sentiment": sentiment["label"],
            "intensity_score": sentiment["intensity_score"],
        })
    scored_reviews.sort(key=lambda x: x["credibility_score"], reverse=True)

    dim_summary = compute_dimension_summary(review_items)
    radar = build_radar_data(dim_summary)
    leads = generate_leads(review_items, dim_summary, scored_reviews)
    anomalies = detect_anomaly_patterns(dim_summary)

    avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else None

    # Overall risk assessment
    critical_count = sum(1 for l in leads if l["severity"] == "critical")
    high_count = sum(1 for l in leads if l["severity"] == "high")
    medium_count = sum(1 for l in leads if l["severity"] == "medium")

    if critical_count >= 1:
        overall_risk = "critical"
    elif high_count >= 2:
        overall_risk = "high"
    elif high_count >= 1 or medium_count >= 3:
        overall_risk = "medium"
    else:
        overall_risk = "low"

    result = {
        "matched": True,
        "name": args.name,
        "school": args.school,
        "college": args.college,
        "review_count": len(all_review_texts),
        "rating_stats": {
            "average": avg_rating,
            "count": len(ratings),
            "distribution": dict(Counter(ratings).most_common()) if ratings else {},
        },
        "credibility": {
            "average_score": round(
                sum(r["credibility_score"] for r in scored_reviews) / len(scored_reviews), 2
            ) if scored_reviews else None,
            "top_reviews": scored_reviews[:args.top_reviews],
        },
        "dimension_summary": dim_summary,
        "radar_data": radar,
        "cross_dimensional_anomalies": anomalies,
        "investigation_leads": leads,
        "overall_risk_assessment": {
            "level": overall_risk,
            "critical_leads": critical_count,
            "high_leads": high_count,
            "medium_leads": medium_count,
            "summary": (
                f"发现 {len(leads)} 条调查线索，其中 critical {critical_count} 条、high {high_count} 条、"
                f"medium {medium_count} 条。整体风险等级: {overall_risk}。"
            ),
        },
        "disclaimer": (
            "本分析基于匿名第三方学生评价数据库。所有线索均为假设生成器，必须通过可验证的公开记录交叉验证后 "
            "才能纳入最终调查报告。评价内容不代表调查者立场。"
        ),
    }

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"[INFO] Result saved to: {args.output}")
    print(f"[SUMMARY] Reviews: {result['review_count']}, "
          f"Avg rating: {result['rating_stats']['average']}, "
          f"Avg credibility: {result['credibility']['average_score']}, "
          f"Leads: {len(leads)}, Anomalies: {len(anomalies)}, Risk: {overall_risk}")
    if leads:
        for lead in leads[:5]:
            print(f"  - [{lead['severity'].upper()}] {lead['label']} (mentions: {lead['mention_count']}, confidence: {lead['confidence_level']})")
    if anomalies:
        for a in anomalies:
            print(f"  - [ANOMALY] {a['pattern']}: {a['description'][:60]}...")


if __name__ == "__main__":
    main()
