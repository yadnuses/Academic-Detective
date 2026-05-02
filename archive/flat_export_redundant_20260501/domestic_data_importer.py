#!/usr/bin/env python3
"""
data_importer.py

Unified importer for academic database exports.
Supports CNKI (.xlsx), Wanfang (.csv), Web of Science (.ris), and JSON.
Outputs a standardized unified_papers.json with deduplication.

Usage:
    python data_importer.py --cnki ./data/cnki.xlsx --output ./data/unified_papers.json
    python data_importer.py --cnki ./data/cnki.xlsx --wanfang ./data/wanfang.csv --output ./data/unified_papers.json --deduplicate
"""

import csv
import json
import argparse
import difflib
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from utils import get_logger, save_json
from db import InvestigationDB

logger = get_logger("data_importer")


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_cnki_xlsx(path: Path) -> list[dict]:
    """Parse CNKI export Excel file."""
    if not HAS_OPENPYXL:
        logger.error("openpyxl is required for .xlsx parsing. Install: pip install openpyxl")
        return []

    papers = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        # CNKI typically has headers in row 1
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Map common CNKI headers
        title_idx = _find_col(headers, ["题名", "Title", "标题", "文章标题"])
        author_idx = _find_col(headers, ["作者", "Author", "Authors", "责任人"])
        journal_idx = _find_col(headers, ["来源", "Source", "期刊", "刊名"])
        year_idx = _find_col(headers, ["年", "Year", "发表时间", "PubTime"])
        doi_idx = _find_col(headers, ["DOI", "Doi"])

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            paper = {
                "title": _cell_str(row, title_idx),
                "authors": _split_authors(_cell_str(row, author_idx)),
                "journal": _cell_str(row, journal_idx),
                "year": _parse_year(_cell_str(row, year_idx)),
                "doi": _cell_str(row, doi_idx) if doi_idx >= 0 else None,
                "source_db": "cnki",
                "raw_export_row": {h: str(v) if v is not None else "" for h, v in zip(headers, row)},
            }
            if paper["title"]:
                papers.append(paper)
    except Exception as e:
        logger.error("Failed to parse CNKI xlsx: %s", e)
    return papers


def parse_wanfang_csv(path: Path) -> list[dict]:
    """Parse Wanfang export CSV file."""
    papers = []
    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                paper = {
                    "title": row.get("Title", row.get("题名", "")).strip(),
                    "authors": _split_authors(row.get("Author", row.get("作者", ""))),
                    "journal": row.get("Source", row.get("刊名", "")).strip() or None,
                    "year": _parse_year(row.get("Year", row.get("年", ""))),
                    "doi": row.get("DOI", "").strip() or None,
                    "source_db": "wanfang",
                    "raw_export_row": dict(row),
                }
                if paper["title"]:
                    papers.append(paper)
    except Exception as e:
        logger.error("Failed to parse Wanfang csv: %s", e)
    return papers


def parse_wos_ris(path: Path) -> list[dict]:
    """Parse Web of Science RIS file (simplified)."""
    papers = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            current = {}
            for line in f:
                line = line.strip()
                if not line:
                    continue
                if line.startswith("TY  -"):
                    if current.get("title"):
                        papers.append(_normalize_ris_record(current))
                    current = {"source_db": "wos"}
                elif line.startswith("TI  -"):
                    current["title"] = line[5:].strip()
                elif line.startswith("AU  -"):
                    current.setdefault("authors", []).append(line[5:].strip())
                elif line.startswith("SO  -"):
                    current["journal"] = line[5:].strip()
                elif line.startswith("PY  -"):
                    current["year"] = _parse_year(line[5:].strip())
                elif line.startswith("DO  -"):
                    current["doi"] = line[5:].strip()
            if current.get("title"):
                papers.append(_normalize_ris_record(current))
    except Exception as e:
        logger.error("Failed to parse WoS ris: %s", e)
    return papers


def parse_json(path: Path) -> list[dict]:
    """Parse existing unified_papers.json or similar."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict) and "papers" in data:
            return data["papers"]
        if isinstance(data, list):
            return data
    except Exception as e:
        logger.error("Failed to parse JSON: %s", e)
    return []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_col(headers: list[str], candidates: list[str]) -> int:
    for cand in candidates:
        for i, h in enumerate(headers):
            if cand in h or h in cand:
                return i
    return -1


def _cell_str(row: tuple, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _split_authors(author_str: str) -> list[str]:
    if not author_str:
        return []
    # Handle "张三;李四;王五" or "张三,李四,王五"
    for sep in [";", ",", "，"]:
        if sep in author_str:
            return [a.strip() for a in author_str.split(sep) if a.strip()]
    return [author_str.strip()]


def _parse_year(val) -> Optional[int]:
    if val is None:
        return None
    s = str(val).strip()
    # Extract 4-digit year
    for part in s.split():
        if part.isdigit() and len(part) == 4:
            year = int(part)
            if 1950 <= year <= 2030:
                return year
    return None


def _normalize_ris_record(record: dict) -> dict:
    return {
        "title": record.get("title", ""),
        "authors": record.get("authors", []),
        "journal": record.get("journal"),
        "year": record.get("year"),
        "doi": record.get("doi"),
        "source_db": record.get("source_db", "wos"),
        "raw_export_row": {},
    }


def _normalize_title(title: str) -> str:
    """Normalize title for deduplication."""
    return title.lower().strip().replace(" ", "").replace("\u3000", "")


def _title_similarity(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, _normalize_title(a), _normalize_title(b)).ratio()


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def deduplicate(papers: list[dict], threshold: float = 0.85) -> tuple[list[dict], int]:
    """
    Remove duplicates from paper list.

    Strategy:
        1. Exact match: title + year + first_author
        2. Fuzzy match: title similarity >= threshold

    Returns:
        (deduplicated_papers, duplicates_removed_count)
    """
    unique = []
    removed = 0

    for p in papers:
        title = p.get("title", "")
        year = p.get("year")
        authors = p.get("authors", [])
        first_author = authors[0] if authors else ""

        # Check exact duplicate
        is_dup = False
        for u in unique:
            if (u.get("year") == year
                and u.get("authors", [""])[0:1] == [first_author]
                and _title_similarity(u.get("title", ""), title) >= 0.98):
                is_dup = True
                break
            # Fuzzy match
            if _title_similarity(u.get("title", ""), title) >= threshold:
                is_dup = True
                break

        if is_dup:
            removed += 1
            logger.debug("Removed duplicate: %s", title[:60])
        else:
            unique.append(p)

    return unique, removed


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_import(args) -> dict:
    all_papers = []
    sources = []

    if args.cnki:
        papers = parse_cnki_xlsx(Path(args.cnki))
        all_papers.extend(papers)
        sources.append(("cnki", args.cnki, len(papers)))
        logger.info("CNKI: imported %d papers from %s", len(papers), args.cnki)

    if args.wanfang:
        papers = parse_wanfang_csv(Path(args.wanfang))
        all_papers.extend(papers)
        sources.append(("wanfang", args.wanfang, len(papers)))
        logger.info("Wanfang: imported %d papers from %s", len(papers), args.wanfang)

    if args.wos:
        papers = parse_wos_ris(Path(args.wos))
        all_papers.extend(papers)
        sources.append(("wos", args.wos, len(papers)))
        logger.info("WoS: imported %d papers from %s", len(papers), args.wos)

    if args.json:
        papers = parse_json(Path(args.json))
        all_papers.extend(papers)
        sources.append(("json", args.json, len(papers)))
        logger.info("JSON: imported %d papers from %s", len(papers), args.json)

    if not all_papers:
        logger.error("No papers imported. Provide at least one input file.")
        return {}

    # Deduplicate
    duplicates_removed = 0
    if args.deduplicate:
        all_papers, duplicates_removed = deduplicate(all_papers, threshold=args.threshold)
        logger.info("Deduplication: removed %d duplicates", duplicates_removed)

    result = {
        "papers": all_papers,
        "metadata": {
            "import_date": datetime.now().isoformat(),
            "source_files": [{"type": s[0], "path": s[1], "records": s[2]} for s in sources],
            "total_papers": len(all_papers),
            "duplicates_removed": duplicates_removed,
        },
    }

    # Save JSON
    output_path = Path(args.output)
    save_json(result, output_path)
    logger.info("Saved unified papers to %s (%d papers)", output_path, len(all_papers))

    # Log to SQLite if case-dir provided
    if args.case_dir:
        db = InvestigationDB(args.case_dir)
        db.init_schema()
        for s in sources:
            db.log_import(
                case_name=Path(args.case_dir).name,
                source_type=s[0],
                source_file=s[1],
                records_imported=s[2],
                duplicates_removed=duplicates_removed,
            )
        logger.info("Import logged to SQLite")

    return result


def main():
    parser = argparse.ArgumentParser(description="Unified academic database importer")
    parser.add_argument("--case-dir", "-C", help="Case working directory (for SQLite logging)")
    parser.add_argument("--cnki", help="Path to CNKI export .xlsx")
    parser.add_argument("--wanfang", help="Path to Wanfang export .csv")
    parser.add_argument("--wos", help="Path to Web of Science export .ris")
    parser.add_argument("--json", "-j", help="Path to existing unified_papers.json")
    parser.add_argument("--output", "-o", required=True, help="Output path for unified_papers.json")
    parser.add_argument("--deduplicate", "-d", action="store_true", help="Enable deduplication")
    parser.add_argument("--threshold", type=float, default=0.85, help="Fuzzy dedup threshold (0-1)")
    args = parser.parse_args()

    run_import(args)


if __name__ == "__main__":
    main()
